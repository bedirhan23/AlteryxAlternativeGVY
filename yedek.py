import re
import glob
import os
import shutil
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from tkinter import messagebox
import locale
from unidecode import unidecode

locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')

class PdfExtractor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.extracted_text = None

    def extract_text_from_pdf(self, pdf_file):
        try:
            reader = PdfReader(pdf_file)
            self.extracted_text = ""

            for page in reader.pages:
                self.extracted_text += page.extract_text()

            return self.extracted_text
        except Exception as e:
            print(f"Hata oluştu: {e}")

    def get_extracted_text(self):
        extracted_text = self.extracted_text
        icra_dairesi_match = re.search('(.+) İcra Dairesi', extracted_text)
        self.icra_dairesi = icra_dairesi_match.group(1) if icra_dairesi_match else None

        icra_dosyasi_match = re.search('(.+) İcra Dosyası', extracted_text)
        self.icra_dosyasi = icra_dosyasi_match.group(1) if icra_dosyasi_match else None

        borclu_tckn = re.search(r'Borçlu\s*:\s*([^\/]+)\s*\/\s*(\d{11})', extracted_text)
        self.borclu = borclu_tckn.group(1) if borclu_tckn else None
        self.tckn = borclu_tckn.group(2) if borclu_tckn else None
        self.birlikte = borclu_tckn if borclu_tckn else None

        alacak_match = re.findall(r'([\d.,]+) TL', extracted_text)
        self.alacak = alacak_match[0] if alacak_match else None

        feragat_match = re.findall(r'([\d.,]+) TL', extracted_text)
        self.feragat = feragat_match[1] if len(feragat_match) >= 2 else None

        url_match = re.search(r'https://evrakdogrula\.uyap\.gov\.tr/[a-zA-Z0-9]+', extracted_text)
        self.url = url_match.group() if url_match else None

class ConverterFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)

        options = {'padx': 5, 'pady': 5}

        self.output_folder_path = None
        self.output_xlsx_file = None

        self.input_button = ttk.Button(self, text='Input Folder', command=self.call_folder, state='disabled')
        self.input_button.grid(column=0, row=2, **options)

        self.output_button = ttk.Button(self, text='Output Folder', command=self.call_output_folder, state='normal')
        self.output_button.grid(column=0, row=0, **options)

        self.output_txt_button = ttk.Button(self, text='Output Excel', command=self.call_output_xlsx, state='disabled')
        self.output_txt_button.grid(column=0, row=1, **options)

        self.summary_button = ttk.Button(self, text='Ozet', command=self.show_summary, state='normal')
        self.summary_button.grid(column=0, row= 4, **options)
        self.information_label = ttk.Label(self, text='Butonlar sırasıyla çalışmaktadır.')
        self.information_label.grid(column=0, row=5, **options)

    def show_summary(self):
        summary_text = f'{self.outlier_counter} dosya formata uymadığı {self.output_folder_path} klasörüne aktarıldı.\n {self.filled_row} dosya {self.output_xlsx_file} dosyasına işlendi'
        messagebox.showinfo('Özet', summary_text)

    def call_folder(self):
        folder_path = fd.askdirectory()
        if folder_path:
            self.process_folder(folder_path)
            messagebox.showinfo('Başarılı', 'Klasör başarıyla seçildi!')
            self.input_button['state'] = 'disabled'

    def call_output_folder(self):
        self.output_folder_path = fd.askdirectory()
        if self.output_folder_path:
            self.process_folder(self.output_folder_path)
            messagebox.showinfo('Başarılı', "Uyumsuz PDF'lerin klasörü başarıyla seçildi")
            self.output_txt_button['state'] = 'normal'
            self.output_button['state'] = 'disabled'

    def call_output_xlsx(self):
        output_xlsx_file = fd.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_xlsx_file:
            self.output_xlsx_file = output_xlsx_file
            self.process_folder(self.output_xlsx_file)
            messagebox.showinfo('Başarılı', "Verilerin aktarılacağı Excel başarıyla seçildi")
            self.output_txt_button['state'] = 'disabled'
            self.input_button['state'] = 'normal'

    def process_folder(self, folder_path):
        if self.output_xlsx_file is None:
            print("Lütfen önce Output Txt dosyası seçin.")
            return
        if self.output_folder_path is None:
            print("Lütfen önce Output Folder'ı seçin.")
            return

        pdf_extractor = PdfExtractor(folder_path)

        workbook = load_workbook(filename=self.output_xlsx_file)
        self.sheet = workbook.active

        self.outlier_counter = 0
        self.filled_row = 0
        self.tckn_rows = []
        self.tckn_row_mapping = {}

        for pdf_file in glob.glob(os.path.join(folder_path, "*.pdf")):
            pdf_extractor.extract_text_from_pdf(pdf_file)
            pdf_extractor.get_extracted_text()

            self.tckn_cell = None
            found_match = False

            if pdf_extractor.borclu is None or pdf_extractor.feragat is None or len(
                    pdf_extractor.alacak) < 2 or pdf_extractor.tckn is None:
                destination_path = os.path.join(self.output_folder_path, os.path.basename(pdf_file))
                print(f"Moving file {pdf_file} to {self.output_folder_path}")
                shutil.move(pdf_file, destination_path)
                self.outlier_counter += 1
                continue

            print(f"File: {pdf_file}")
            print("İcra Dairesi:", pdf_extractor.icra_dairesi)
            print("İcra Dosyası:", pdf_extractor.icra_dosyasi)
            print("Borçlu:", pdf_extractor.borclu)
            print("TCKN:", pdf_extractor.tckn)
            print(type(pdf_extractor.tckn))
            print("Asıl Alacak:", pdf_extractor.alacak)
            print("Feragat Edilen Tutar:", pdf_extractor.feragat)
            print("URL:", pdf_extractor.url)
            print("********************************")
            print("\n")


            # Check if icra_dosyasi and icra_dairesi match with the values in the 9th and 10th columns
            for current_row in self.sheet.iter_rows(min_row=1, max_row=self.sheet.max_row):
                if current_row[9].value and current_row[10].value:
                    # Convert both strings to their closest ASCII representation
                    excel_icra_dosyasi = unidecode(str(current_row[8].value))
                    excel_icra_dairesi = unidecode(str(current_row[9].value))
                    excel_tckn_value = unidecode(str(current_row[2].value))

                    if unidecode(pdf_extractor.icra_dosyasi.lower()) == excel_icra_dosyasi.lower() and unidecode(
                            pdf_extractor.icra_dairesi.lower()) == excel_icra_dairesi.lower():
                        # If icra_dosyasi and icra_dairesi match, check for TCKN in the same row
                        try:
                            excel_tckn_value_cell = current_row[tckn_column_index]
                            if excel_tckn_value_cell is not None and excel_tckn_value_cell.value is not None:
                                self.tckn_cell = current_row[0]
                                self.sheet.cell(row=current_row[0].row, column=14).value = pdf_extractor.alacak
                                self.sheet.cell(row=current_row[0].row, column=13).value = pdf_extractor.feragat
                                self.sheet.cell(row=current_row[0].row, column=17).value = pdf_extractor.url
                                print(f"icra no ve daire eşit excel'e doldurdum icra dosyası: {excel_icra_dosyasi} and icra dairesi: {excel_icra_dairesi} ", excel_tckn_value_cell)
                                self.filled_row += 1
                                found_match = True
                                break  # Exit the loop when a match is found
                        except IndexError:
                            # Handle error when index is out of range
                            print(f"IndexError: TCKN column might be missing.")
                            continue

                    else:
                        #print("No match found. Checking TCKN.")
                        try:
                            # print(f"Current row: {current_row}")
                            # print(f"ICRA Dosyasi: {excel_icra_dosyasi}")
                            # print(f"ICRA Dairesi: {excel_icra_dairesi}")
                            # print(f"TCKNN DECODE: {excel_tckn_value}")
                            # print("++++++++++++++++++++++++")
                            ...

                            # Identify the column index where TCKN is expected in the Excel sheet
                            tckn_column_index = 2  # Adjust this index based on the actual column index for TCKN

                            # Extract TCKN value from the current_row, applying unidecode
                            excel_tckn_value_cell = current_row[tckn_column_index]

                            # Check if the cell is not empty before extracting the value
                            if excel_tckn_value_cell is not None and excel_tckn_value_cell.value is not None:
                                excel_tckn_value = unidecode(str(excel_tckn_value_cell.value))
                                int_excel_tckn_value = int(excel_tckn_value)
                                self.tckn_int = int(pdf_extractor.tckn)
                                #print("Converted to int:", self.tckn_int)

                                # print(f" 198 Excel TCKN VALUE {excel_tckn_value}, type {type(excel_tckn_value)}")
                                # print(f" 199 SELF TCKN INT {self.tckn_int}, type {type(self.tckn_int)}")

                                if int_excel_tckn_value == self.tckn_int:
                                    print("Match found. Appending to TCKN rows.")
                                    self.tckn_rows.append(self.tckn_int)

                                    self.tckn_row_mapping[self.tckn_int] =current_row[2].row
                                #else:
                                    # print("Condition not met. Values:")
                                    # print("TCKN CELL VALUE:", excel_tckn_value)
                                    # print("PDF TCKN VALUE:", pdf_extractor.tckn)
                            else:
                                print("TCKN column is empty in the current row.")
                        except ValueError as ve:
                            print(f"Error converting to int: {ve}")

            print("TCKN ROWS: ", self.tckn_rows)
            if not found_match:
                if self.tckn_rows:
                    print("ZORLUUUUUUUUU")
                    if len(self.tckn_rows) > 1:
                        destination_path = os.path.join(self.output_folder_path, os.path.basename(pdf_file))
                        print(f"Moving file {pdf_file} to {self.output_folder_path}")
                        shutil.move(pdf_file, destination_path)
                        self.outlier_counter += 1
                        continue

                    for tckn in self.tckn_rows:
                        try:
                            row_number = self.tckn_row_mapping.get(tckn)
                            if row_number is not None:
                                self.sheet.cell(row=row_number, column=14).value = pdf_extractor.alacak
                                self.sheet.cell(row=row_number, column=13).value = pdf_extractor.feragat
                                self.sheet.cell(row=row_number, column=17).value = pdf_extractor.url
                                print("tckn'ye göre doldurdum", )
                                self.filled_row += 1
                        except IndexError:
                            print(f"IndexError: TCKN column might be missing.")

        workbook.save(self.output_xlsx_file)


root = tk.Tk()
root.geometry('200x200')
app = ConverterFrame(root)
app.grid(row=0, column=0, padx=10, pady=10)
root.mainloop()
