import re
import glob
import os
import shutil
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from tkinter import messagebox
import locale
from unidecode import unidecode
import tkinter.font
from tkinter import font

#tk.Tk.option_add("*Font", "Arial 12")
locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')


class PdfExtractor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.extracted_text = None

    def extract_text_from_pdf(self, pdf_file):
        try:
            reader = PdfReader(pdf_file)
            self.extracted_text = ""

            for page in reader.pages:
                self.extracted_text += page.extract_text()

            return self.extracted_text
        except Exception as e:
            print(f"Hata oluştu: {e}")

    def get_extracted_text(self):
        extracted_text = self.extracted_text
        icra_dairesi_match = re.search('(.+) İcra Dairesi', extracted_text)
        self.icra_dairesi = icra_dairesi_match.group(1) if icra_dairesi_match else None

        icra_dosyasi_match = re.search('(.+) İcra Dosyası', extracted_text)
        self.icra_dosyasi = icra_dosyasi_match.group(1) if icra_dosyasi_match else None

        borclu_tckn = re.search(r'Borçlu\s*:\s*([^\/]+)\s*\/\s*(\d{11})', extracted_text)
        self.borclu = borclu_tckn.group(1) if borclu_tckn else None
        self.tckn = borclu_tckn.group(2) if borclu_tckn else None
        self.birlikte = borclu_tckn if borclu_tckn else None

        alacak_match = re.findall(r'([\d.,]+) TL', extracted_text)
        self.alacak = alacak_match[0] if alacak_match else None

        feragat_match = re.findall(r'([\d.,]+) TL', extracted_text)
        self.feragat = feragat_match[1] if len(feragat_match) >= 2 else None

        url_match = re.search(r'https://evrakdogrula\.uyap\.gov\.tr/[a-zA-Z0-9]+', extracted_text)
        self.url = url_match.group() if url_match else None


class ConverterFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)

        options = {'padx': 5, 'pady': 5}

        self.desired_font = tkinter.font.Font(family="Arial", size=16)

        self.output_folder_path = None
        self.output_xlsx_file = None
        self.empty_excel_file = None

        self.set_button_style()

        self.input_button = ttk.Button(self, text="PDF'lerin Klasörü", command=self.call_folder, state='disabled')
        self.input_button.grid(column=0, row=3, padx=50, pady=20)
        self.input_button.configure(style="Bold.TButton")  # Correct font configuration

        # Correctly configure other buttons
        self.output_button = ttk.Button(self, text='İstenmeyen Klasörü', command=self.call_output_folder, state='normal')
        self.output_button.grid(column=0, row=0, padx=50, pady=20)
        self.output_button.configure(style="Bold.TButton")

        self.output_txt_button = ttk.Button(self, text='Doldurulacak  Excel', command=self.call_output_xlsx, state='disabled')
        self.output_txt_button.grid(column=0, row=1, padx=50, pady=20)
        self.output_txt_button.configure(style="Bold.TButton")

        self.empty_excel_button = ttk.Button(self, text='Boş Excel', command=self.empty_excel, state='disabled')
        self.empty_excel_button.grid(column=0, row=2, padx=50, pady=20)
        self.empty_excel_button.configure(style="Bold.TButton")

        self.summary_button = ttk.Button(self, text='Ozet', command=self.show_summary, state='normal')
        self.summary_button.grid(column=0, row=4, padx=50, pady=20)
        self.summary_button.configure(style="Bold.TButton")

        self.information_label = ttk.Label(self, text='Butonlar sırasıyla çalışmaktadır.')
        self.information_label.grid(column=0, row=5, padx=50, pady=20)
        self.information_label.configure(font= self.desired_font)


    def set_button_style(self):
        style = ttk.Style()
        style.configure("Bold.TButton", font= self.desired_font)

    def show_summary(self):
        summary_text = f'{self.outlier_counter} dosya formata uymadığı {self.output_folder_path} klasörüne aktarıldı.\n {self.filled_row} dosya {self.output_xlsx_file} dosyasına işlendi'
        messagebox.showinfo('Özet', summary_text)

    def call_folder(self):
        folder_path = fd.askdirectory()
        if folder_path:
            self.process_folder(folder_path)
            messagebox.showinfo('Başarılı', 'Klasör başarıyla seçildi!')
            self.input_button['state'] = 'disabled'

    def call_output_folder(self):
        self.output_folder_path = fd.askdirectory()
        if self.output_folder_path:
            self.process_folder(self.output_folder_path)
            messagebox.showinfo('Başarılı', "Uyumsuz PDF'lerin klasörü başarıyla seçildi")
            self.output_txt_button['state'] = 'normal'
            self.output_button['state'] = 'disabled'

    def call_output_xlsx(self):
        output_xlsx_file = fd.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_xlsx_file:
            self.output_xlsx_file = output_xlsx_file
            self.empty_excel_button['state'] = 'normal'  # Enable empty_excel_button when output_xlsx_file is selected
            messagebox.showinfo('Başarılı', "Verilerin aktarılacağı Excel başarıyla seçildi")
            self.output_txt_button['state'] = 'disabled'
            self.input_button['state'] = 'normal'

    def empty_excel(self):
        print("Empty Excel dosyası seçiliyor...")
        self.empty_excel_file = fd.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if self.empty_excel_file:
            messagebox.showinfo('Başarılı', message="Boş excel dosyası başarıyla seçildi")
            self.empty_excel_button['state'] = 'disabled'

    def process_folder(self, folder_path):
        if self.output_xlsx_file is None:
            print("Lütfen önce Output Txt dosyası seçin.")
            return
        if self.output_folder_path is None:
            print("Lütfen önce Output Folder'ı seçin.")
            return

        if self.empty_excel_file is None:
            print("Lütfen önce Boş Excel dosyası seçin.")
            return

        pdf_extractor = PdfExtractor(folder_path)
        workbook = load_workbook(filename=self.output_xlsx_file)
        self.sheet = workbook.active

        workbook2 = load_workbook(filename=self.empty_excel_file)
        self.empty_sheet = workbook2.active

        self.outlier_counter = 0
        self.filled_row = 0
        self.tckn_rows = []
        read_tckns = []
        self.tckn_row_mapping = []
        filled = []
        self.empty_row = 2

        excel_tckn = []
        excel_icra = []
        for a in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row):
            excel_tckn.append(unidecode(str(a[2].value)))
            excel_icra.append(unidecode(str(a[8].value).lower() +"_"+ str(a[9].value).lower()))

        # Initialize tckn_column_index outside the loop
        tckn_column_index = 2  # Adjust this index based on the actual column index for TCKN

        # Step 1: Iterate through PDFs and fill rows based on icra_dairesi and icra_dosyasi
        for pdf_file in glob.glob(os.path.join(folder_path, "*.pdf")):
            pdf_extractor.extract_text_from_pdf(pdf_file)
            pdf_extractor.get_extracted_text()

            print(f"File: {pdf_file}")
            self.empty_sheet.cell(row= self.empty_row, column=1).value = pdf_file
            print("İcra Dairesi:", pdf_extractor.icra_dairesi)
            self.empty_sheet.cell(row=self.empty_row, column=2).value = pdf_extractor.icra_dairesi
            print("İcra Dosyası:", pdf_extractor.icra_dosyasi)
            self.empty_sheet.cell(row=self.empty_row, column=3).value = pdf_extractor.icra_dosyasi
            print("Borçlu:", pdf_extractor.borclu)
            self.empty_sheet.cell(row=self.empty_row, column=4).value = pdf_extractor.borclu
            print("TCKN:", pdf_extractor.tckn)
            self.empty_sheet.cell(row=self.empty_row, column=5).value = pdf_extractor.tckn
            print(type(pdf_extractor.tckn))
            print("Asıl Alacak:", pdf_extractor.alacak)
            self.empty_sheet.cell(row=self.empty_row, column=6).value = pdf_extractor.alacak
            print("Feragat Edilen Tutar:", pdf_extractor.feragat)
            self.empty_sheet.cell(row=self.empty_row, column=8).value = pdf_extractor.feragat
            print("URL:", pdf_extractor.url)
            self.empty_sheet.cell(row=self.empty_row, column=9).value = pdf_extractor.url
            print("********************************")
            print("\n")

            self.empty_row += 1
            self.tckn_cell = None
            found_match = False

            # Check if essential information is missing in the PDF
            if pdf_extractor.borclu is None or pdf_extractor.feragat is None or len(pdf_extractor.alacak) < 2 or pdf_extractor.tckn is None:
                destination_path = os.path.join(self.output_folder_path, "Undesired_Files")
                os.makedirs(destination_path, exist_ok=True)

                if os.path.exists(pdf_file):
                    try:
                        shutil.move(pdf_file, os.path.join(destination_path, os.path.basename(pdf_file)))
                        self.outlier_counter += 1
                    except FileNotFoundError as e:
                        print(f"Error moving file: {e}")
                else:
                    print(f"File not found: {pdf_file}")

                continue

            try:
                icra_row = excel_icra.index(unidecode(pdf_extractor.icra_dosyasi.lower() +"_"+ pdf_extractor.icra_dairesi.lower())) + 2
                found_match = True

                if self.sheet.cell(row=icra_row, column=13).value is not None and self.sheet.cell(row= icra_row, column=14).value is not None:
                    print(f"Warning: Row {icra_row}, columns 13 and 14 are already filled. Skipping...")
                else:
                    self.sheet.cell(row=icra_row, column=14).value = pdf_extractor.alacak
                    self.sheet.cell(row=icra_row, column=13).value = pdf_extractor.feragat
                    self.sheet.cell(row=icra_row, column=17).value = pdf_extractor.url
                    self.filled_row += 1
                    print(f"icra no ve daire eşit excel'e doldurdum icra dosyası: {excel_icra[icra_row - 2]}")


            except ValueError as ve:
                pass

            # Step 2: Check if the TCKN is unique in the Excel file
            #unique_tckns = set([str(row[2].value) for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row)])

            current_pdf_tckn = str(pdf_extractor.tckn)

            if found_match:
                continue


            if excel_tckn.count(current_pdf_tckn) == 1:

                current_row = excel_tckn.index(current_pdf_tckn) + 2
                self.sheet.cell(row=current_row, column=14).value = pdf_extractor.alacak
                self.sheet.cell(row=current_row, column=13).value = pdf_extractor.feragat
                self.sheet.cell(row=current_row, column=17).value = pdf_extractor.url
                self.sheet.cell(row=current_row, column=11).value = pdf_extractor.icra_dosyasi
                self.sheet.cell(row=current_row, column=12).value = pdf_extractor.icra_dairesi
                self.tckn_rows.append(int(current_pdf_tckn))
                self.filled_row += 1
                print(f"PDF TCKN found and updated in Excel: {current_pdf_tckn}")

            else:
                destination_path = os.path.join(self.output_folder_path, "Undesired_Files")
                os.makedirs(destination_path, exist_ok=True)
                print(f"Moving file {pdf_file} to {destination_path}")
                if os.path.exists(pdf_file):
                    try:
                        shutil.move(pdf_file,
                                    os.path.normpath(os.path.join(destination_path, os.path.basename(pdf_file))))
                        self.outlier_counter += 1
                    except FileNotFoundError as e:
                        print(f"Error moving file: {e}")
                else:
                    print(f"File not found: {pdf_file}")

        workbook2.save(self.empty_excel_file)
        workbook.save(self.output_xlsx_file)

root = tk.Tk()
#root.configure(font = self.desired_font)
root.wm_minsize(width=400, height=300)
app = ConverterFrame(root)
app.grid(row=0, column=0, padx=10, pady=10)
root.mainloop()
